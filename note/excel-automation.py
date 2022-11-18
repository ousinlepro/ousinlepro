from turtle import heading
from openpyxl import load_workbook
from openpyxl.workbook import workbook
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from subprocess import Popen
#from moyenne import *

wb = load_workbook('note.xlsx')

ws = wb.active

ws['A7'] = 'TEST'
cells = ws['A9:A11']
cells = 'salut'
#cell.font = Font(name='Arial',bold=True,color='00339966')

#ws.append(['HG','ECO-FAM', 'PC', 'EC', 'SVT', 'EPS', 'ESPAGNOL','MATHS', 'ANGLAIS','INFORMATIQUE', 'DICTEE', 'TSQ', 'REDACTON'])

"""
for row in range(1,11):
    for col in range(1,5):
        char = get_column_letter(col)
        ws[char + str(row)].value = char + str(row) """

data = {
    "ousin":{
        "hg":19,
        "svt":16,
        "pc": 20
    },
    "Ameth":{
        'hg':15,
        "svt":12,
        "pc":11
    }
}

# Save the file
wb.save("note.xlsx")
Popen('start note.xlsx',shell=True)
#print(f'{ws["A1"].value}: {ws["B1"].value}')